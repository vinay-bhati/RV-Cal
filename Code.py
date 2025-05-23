import streamlit as st
import pandas as pd
import numpy as np
import math
from validate_email import validate_email
import boto3
from io import StringIO,BytesIO
from datetime import date, datetime
import io
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# Load AWS configuration from secrets
access_key = st.secrets["aws"]["access_key"]
secret_key = st.secrets["aws"]["secret_key"]
bucket_name = st.secrets["aws"]["bucket_name"]
s3_filename = st.secrets["aws"]["s3_filename"]

# # AWS 
# access_key = os.environ["AWS_ACCESS_KEY_ID"]
# secret_key = os.environ["AWS_SECRET_ACCESS_KEY"]
# bucket_name = os.environ["BUCKET_NAME"]
# s3_filename = os.environ["S3_FILENAME"]

# Initialize the S3 client
s3_client = boto3.client(
    's3',
    aws_access_key_id=access_key,
    aws_secret_access_key=secret_key
)
rv_threshold = None
def append_to_s3(email, rv_threshold, standard, has_fvc_pred, gender, age, height, measured_fev1, measured_fvc, fvc_percent_predicted, predicted_fev1, predicted_fvc, predicted_fev1_fvc, rv_percent_est, rv150_prob, race):
    # Prepare the CSV row as a string, note the newline at the start
    date_today = date.today().isoformat()
    date_time = datetime.now().isoformat()
    data_row = f"\n{email}|{rv_threshold}|{standard}|{has_fvc_pred}|{gender}|{age}|{height}|{measured_fev1}|{measured_fvc}|{fvc_percent_predicted}|{round(predicted_fev1, 2) if has_fvc_pred == 'No' else ''}|{round(predicted_fvc, 2) if has_fvc_pred == 'No' else ''}|{round(predicted_fev1_fvc, 2) if has_fvc_pred == 'No' else ''}|{rv_percent_est}|{round(rv150_prob,2)}|{race}|{date_today}|{date_time}"

    # Bucket and file details
    bucket_name = st.secrets["aws"]["bucket_name"]
    s3_filename = st.secrets["aws"]["s3_filename"]

    # Initialize S3 client
    s3_client = boto3.client('s3', aws_access_key_id=st.secrets["aws"]["access_key"], aws_secret_access_key=st.secrets["aws"]["secret_key"])
    # s3_client = boto3.client('s3', aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"], aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"])
    
    # Download the existing CSV from S3
    response = s3_client.get_object(Bucket=bucket_name, Key=s3_filename)
    existing_data = response['Body'].read().decode('utf-8')

    # Append the new data
    updated_data = existing_data + data_row

    # Upload the updated data back to S3
    s3_client.put_object(Bucket=bucket_name, Key=s3_filename, Body=updated_data.encode('utf-8'))

# Initialize session state variables if they don't exist
if 'standard' not in st.session_state:
    st.session_state['standard'] = None
if 'has_fvc_pred' not in st.session_state:
    st.session_state['has_fvc_pred'] = None

# Load data based on gender
def load_data(gender):
    if gender == 'Male':
        df_fev1 = pd.read_csv("Male FEV1.csv")
        df_fvc = pd.read_csv("Male FVC.csv")
        df_fev1_fvc = pd.read_csv("Male FEV1 FVC.csv")
        #df_rv = "Male RV.csv"
    else:
        df_fev1 = pd.read_csv("Female FEV1.csv")
        df_fvc = pd.read_csv("Female FCV.csv")
        df_fev1_fvc = pd.read_csv("Female FEV1 FVC.csv")
        #df_rv = "Female RV.csv"
    return df_fev1, df_fvc, df_fev1_fvc

# Calculate FEV1, FVC, and FEV1/FVC based on gender
def calculate_values(age, height, gender):
    df_fev1, df_fvc, df_fev1_fvc = load_data(gender)
    spline_fev1 = df_fev1[df_fev1['age'] == age].iloc[0]
    spline_fvc = df_fvc[df_fvc['age'] == age].iloc[0]
    spline_fev1_fvc = df_fev1_fvc[df_fev1_fvc['age'] == age].iloc[0]
    #spline_rv = df_rv[df_rv['age'] == age].iloc[0]

    
    if gender == 'Male':
        # Male calculations
        M_FEV1 = np.exp(-11.399108 + 2.462664 * np.log(height) - 0.011394 * np.log(age) + spline_fev1['M Spline'])
        M_FVC = np.exp(-12.629131 + 2.727421 * np.log(height) + 0.009174 * np.log(age) + spline_fvc['M Spline'])
        M_FEV1_FVC = np.exp(1.022608 - 0.218592 * np.log(height) - 0.027586 * np.log(age) + spline_fev1_fvc['M Spline'])
    else:
        # Female calculations
        M_FEV1 = np.exp(-10.901689 + 2.385928 * np.log(height) - 0.076386 * np.log(age) + spline_fev1['M Spline'])
        M_FVC = np.exp(-12.055901 + 2.621579 * np.log(height) - 0.035975 * np.log(age) + spline_fvc['M Spline'])
        M_FEV1_FVC = np.exp(0.9189568 - 0.1840671 * np.log(height) - 0.0461306 * np.log(age) + spline_fev1_fvc['M Spline'])
        #M_RV = np.exp(-2.50593 + 0.01307 * age + 0.01379 * height + spline_rv['M Spline'])
    
    return M_FEV1, M_FVC, M_FEV1_FVC,#M_RV

def calculate_rv_est(percent_predicted_fvc, measured_fev1_fvc, age, gender):
    # Load constants from secrets
    consts = st.secrets["rv_est_constants"]

#     consts = {
#     "fvc_multiplier": float(os.environ["RV_EST_FVC_MULTIPLIER"]),
#     "fev1_fvc_multiplier": float(os.environ["RV_EST_FEV1_FVC_MULTIPLIER"]),
#     "fvc_sqrt_multiplier": float(os.environ["RV_EST_FVC_SQRT_MULTIPLIER"]),
#     "age_multiplier": float(os.environ["RV_EST_AGE_MULTIPLIER"]),
#     "gender_multiplier": float(os.environ["RV_EST_GENDER_MULTIPLIER"]),
#     "constant": float(os.environ["RV_EST_CONSTANT"]),
# }
    
    # Converting gender to numeric value: 1 for Male, 0 for Female
    gender_numeric = 1 if gender == 'Male' else 0
    
    # Use constants from secrets.toml for calculations
    rv_est = round(
        (percent_predicted_fvc * consts["fvc_multiplier"] +
         round(measured_fev1_fvc, 3) * consts["fev1_fvc_multiplier"] +
         np.sqrt(percent_predicted_fvc) * consts["fvc_sqrt_multiplier"] +
         age * consts["age_multiplier"] +
         gender_numeric * consts["gender_multiplier"] +
         consts["constant"]), 1)
    
    return rv_est

def calculate_rv_predicted(rv_percent_est):
    # Load constants from secrets
    preds = st.secrets["rv_pred_constants"]
    
#     preds = {
#     "rv150_coef": float(os.environ["RV_PRED_150_COEF"]),
#     "rv150_intercept": float(os.environ["RV_PRED_150_INTERCEPT"]),
#     "rv175_coef": float(os.environ["RV_PRED_175_COEF"]),
#     "rv175_intercept": float(os.environ["RV_PRED_175_INTERCEPT"]),
#     "rv200_coef": float(os.environ["RV_PRED_200_COEF"]),
#     "rv200_intercept": float(os.environ["RV_PRED_200_INTERCEPT"]),
# }
    
    # Using the logistic regression model to predict probabilities
    rv150 = round((1 / (1 + math.exp(-1 * (preds["rv150_coef"] + preds["rv150_intercept"] * rv_percent_est)))), 3) * 100
    rv175 = round((1 / (1 + math.exp(-1 * (preds["rv175_coef"] + preds["rv175_intercept"] * rv_percent_est)))), 3) * 100
    rv200 = round((1 / (1 + math.exp(-1 * (preds["rv200_coef"] + preds["rv200_intercept"] * rv_percent_est)))), 3) * 100
    
    return rv150, rv175, rv200

def calculate_ecsc_fvc(age, height, fev1, fvc, gender, race):
    """Calculate ECSC based Predicted FVC based on the formula."""
    if gender == 1 and race == 1:
        return round((0.00064 * age - 0.000269 * (age ** 2) + 0.00018642 * (height ** 2) - 0.1933), 2)
    elif gender == 0 and race == 1:
        return round((0.0187 * age - 0.000382 * (age ** 2) + 0.00014815 * (height ** 2) - 0.356), 2)
    elif gender == 1 and race == 2:
        return round((-0.01821 * age + 0.00016643 * (height ** 2) - 0.1517), 2)
    elif gender == 0 and race == 2:
        return round((0.00536 * age - 0.000265 * (age ** 2) + 0.00013606 * (height ** 2) - 0.3039), 2)
    return "Insufficient data for prediction"

def calculate_ecsc_metrics(age, height, measured_fev1, predicted_fvc, measured_fvc,gender,race):
    """Calculate additional respiratory metrics based on user input and predicted FVC."""
    if measured_fvc and predicted_fvc:
        fvc_percent_predicted = round((measured_fvc / predicted_fvc) * 100, 1)
    else:
        fvc_percent_predicted = None

    # Assuming FEV1/FVC ratio is a straightforward division
    if measured_fev1 and measured_fvc:
        local_ration = measured_fev1 / measured_fvc
        fev1_fvc_ratio =  math.ceil(local_ration * 1000) / 1000
    else:
        fev1_fvc_ratio = None

    # Calculate RV % est based on the provided formula
    if age is not None and measured_fev1 and fvc_percent_predicted and fev1_fvc_ratio:
        rv_percent_est = round((fvc_percent_predicted * 3.46 - round(fev1_fvc_ratio, 3) * 179.8 -
                                np.sqrt(fvc_percent_predicted) * 79.53 - age * 0.98 - gender * 10.88 + 737.06), 1)
    else:
        rv_percent_est = None

    # Calculate probabilities for RV150, RV175, RV200
    if rv_percent_est is not None:
        rv150 = round((1 / (1 + np.exp(-1 * (-9.218401 + 0.0572793 * rv_percent_est)))), 3) * 100
        rv175 = round((1 / (1 + np.exp(-1 * (-9.995177 + 0.0551463 * rv_percent_est)))), 3) * 100
        rv200 = round((1 / (1 + np.exp(-1 * (-11.32753 + 0.0561363 * rv_percent_est)))), 3) * 100
    else:
        rv150, rv175, rv200 = None, None, None

    return fvc_percent_predicted, fev1_fvc_ratio, rv_percent_est, rv150, rv175, rv200

# First command: set page configuration
st.set_page_config(
    page_title='RV Estimate Calculator',
    layout='wide',
    initial_sidebar_state='expanded'
)

## CSS to increase font size for specific elements identified via developer tools
st.markdown("""
<style>
/* Targeting paragraphs within the Markdown container for widget labels */
div[data-testid="stMarkdownContainer"] p {
    font-size: 16px !important;  /* Increase font size of labels */
    font-weight: bold !important;  /* Make text bold */
}

/* Customize input fields to increase font size */
input, select, textarea {
    font-size: 16px !important;  /* Increase font size in input fields */
    height: 50px !important;  /* Make input fields taller for better interaction */
}

/* Enhance the visibility of buttons */
button {
    font-size: 18px !important;
    height: 50px !important;
}
/* Customizing the font size for metric values */
[data-testid="stMetricValue"] > div {
    font-size: 16px !important;  /* Set your desired font size here */
}     
h1 {
    font-size: 30px;  /* Larger font size for headings */
} 
</style>
""", unsafe_allow_html=True)

def process_gli_batch_excel(file):
    # Read the uploaded Excel file into a DataFrame
    try:
        df = pd.read_excel(file, engine='openpyxl')
        df.columns = [col.strip().lower() for col in df.columns]  # Normalize column names
    except Exception as e:
        st.error(f"Failed to read the Excel file: {e}")
        return

    # Initialize list to hold results and counters for successes and errors
    results = []
    success_count = 0
    error_count = 0
    
    # Iterate over DataFrame rows
    for index, row in df.iterrows():
        try:
            # Ensure all required data is present
            if pd.isna(row['age']) or pd.isna(row['gender']) or \
               pd.isna(row['fev1 (l)']) or pd.isna(row['fvc (l)']) or pd.isna(row['fvc % predicted']):
                raise ValueError("Missing data in one or more required fields.")

            email1 = email
            age = int(row['age'])
            gender = row['gender']
            measured_fev1 = float(row['fev1 (l)'])
            measured_fvc = float(row['fvc (l)'])
            fvc_percent_predicted = float(row['fvc % predicted'])
            unique_id = row['unique id']

            measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
            measured_fev1_fvc = round(measured_fev1_fvc, 3)
            
            rv_percent_est = calculate_rv_est(fvc_percent_predicted, measured_fev1_fvc, age, gender)
            rv150, rv175, rv200 = calculate_rv_predicted(rv_percent_est)

            results.append({
                "Unique ID": unique_id,
                "Email": email1,
                "Age": age,
                "Gender": gender,
                "FEV1 (L)": measured_fev1,
                "FVC (L)": measured_fvc,
                "FCV % Predicted": fvc_percent_predicted,
                "FEV1/FVC": measured_fev1_fvc,
                #"rv_percent_est": rv_percent_est,
                "Probability of RV>150": rv150
                #"rv175": rv175,
                #"rv200": rv200
            })
            success_count += 1

        except Exception as e:
            error_count += 1
            st.write(f"Error processing record {index+1}: {e}")

    # Convert results to DataFrame
    results_df = pd.DataFrame(results)
    
    # Display results of processing
    st.success(f"Successfully processed: {success_count} records")
    if error_count > 0:
        st.error(f"Failed to process: {error_count} records")
    
    return results_df

def process_gli_batch_no_fvc_pred(file):
    try:
        df = pd.read_excel(file, engine='openpyxl')
        df.columns = [col.strip().lower() for col in df.columns]  # Normalize column names
    except Exception as e:
        st.error(f"Failed to read the Excel file: {e}")
        return None
    results = []
    success_count = 0
    error_count = 0
        
    for index, row in df.iterrows():
        try:
            # Ensure all required data is present
            if pd.isna(row['age']) or pd.isna(row['gender']) or \
               pd.isna(row['height']) or pd.isna(row['fev1 (l)']) or pd.isna(row['fvc (l)']):
                raise ValueError("Missing data in one or more required fields at row {}".format(index + 1))

            email2 = email
            age = int(row['age'])
            gender = row['gender']
            height = float(row['height'])
            measured_fev1 = float(row['fev1 (l)'])
            measured_fvc = float(row['fvc (l)'])
            unique_id = row["unique id"]

            # Perform the calculations as done in the single entry scenario
            fev1, fvc, fev1_fvc = calculate_values(age, height, gender)
            percent_predicted_fev1 = (measured_fev1 / fev1) * 100
            percent_predicted_fvc = (measured_fvc / fvc) * 100
            measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
            measured_fev1_fvc = round(measured_fev1_fvc, 3)
            
            rv_percent_est = calculate_rv_est(percent_predicted_fvc, measured_fev1_fvc, age, gender)
            rv150, rv175, rv200 = calculate_rv_predicted(rv_percent_est)
            percent_predicted_fvc = round(percent_predicted_fvc,1)
            rv150 = round(rv150,1)
            results.append({
                "Unique ID": unique_id,
                "Email": email2,
                "Age": age,
                "Gender": gender,
                "Height": height,
                "FEV1 (L)": measured_fev1,
                "FVC (L)": measured_fvc,
                "FEV1/FVC": measured_fev1_fvc,
                "FVC % Predicted":percent_predicted_fvc,
                "Probability of RV>150": rv150
                #"rv175": rv175,
                #"rv200": rv200
            })
            success_count += 1

        except Exception as e:
            error_count += 1
            st.error(f"Error processing record {index + 1}: {e}")

    results_df = pd.DataFrame(results)
    st.success(f"Successfully processed: {success_count} records")
    if error_count > 0:
        st.error(f"Failed to process: {error_count} records")

    return results_df

def process_ecsc_batch(file):
    try:
        df = pd.read_excel(file, engine='openpyxl')
        df.columns = [col.strip().lower() for col in df.columns]  # Normalize column names
    except Exception as e:
        st.error(f"Failed to read the Excel file: {e}")
        return None

    results = []
    success_count = 0
    error_count = 0

    for index, row in df.iterrows():
        try:
            # Check for necessary columns
            if 'race' not in df.columns:
                raise ValueError("Column 'Race' is missing from the Excel file.")

            # Validate and normalize input data
            if pd.isna(row['age']) or pd.isna(row['gender']) or \
               pd.isna(row['height']) or pd.isna(row['fev1 (l)']) or \
               pd.isna(row['fvc (l)']) or pd.isna(row['race']):
                raise ValueError("Missing data in one or more required fields.")

            email3 = email
            age = int(row['age'])
            gender = row['gender'].strip().title()  # Normalize gender to handle case variations
            height = float(row['height'])
            measured_fev1 = float(row['fev1 (l)'])
            measured_fvc = float(row['fvc (l)'])
            race = row['race'].strip().title()  # Normalize race to handle case variations
            unique_id = row['unique id']

            # Validate gender and race values
            if gender not in ['Male', 'Female']:
                raise ValueError(f"Invalid gender value: {gender}")
            if race not in ['White', 'Black']:
                raise ValueError(f"Invalid race value: {race}")

            # Map gender and race to numeric values for calculations
            gender = 1 if gender == 'Male' else 0
            race = 1 if race == 'White' else 2

            pred_fvc = calculate_ecsc_fvc(age, height, measured_fev1, measured_fvc, gender, race)
            fvc_percent_predicted, fev1_fvc_ratio, rv_percent_est, rv150, rv175, rv200 = calculate_ecsc_metrics(age, height, measured_fev1, pred_fvc, measured_fvc,gender,race)

            gender = 'Male' if gender == 1 else "Female"
            race = 'White' if race == 1 else "Black"

            results.append({
                "Unique ID": unique_id,
                "Email": email3,
                "Age": age,
                "Gender": gender,
                "Height": height,
                "Race":race,
                "FEV1 (L)": measured_fev1,
                "FVC (L)": measured_fvc,
                "FVC % Predicted": fvc_percent_predicted,
                "FEV1/FVC": fev1_fvc_ratio,
                "Probability of RV>150": rv150
                #"rv175": rv175,
                #"rv200": rv200
            })
            success_count += 1

        except Exception as e:
            error_count += 1
            st.error(f"Error processing record {index + 1}: {e}")

    results_df = pd.DataFrame(results)
    st.success(f"Successfully processed: {success_count} records")
    if error_count > 0:
        st.error(f"Failed to process: {error_count} records")

    return results_df
    
st.title('RV ESTIMATE CALCULATOR')
email = st.text_input("Enter email ID:")
process_type = st.radio("Choose the type of process:", ('Single', 'Batch'),horizontal=True,index=None)

if process_type == 'Single':
    
    if email:
        is_email_valid = validate_email(email)
        if not is_email_valid:
            st.error("Invalid email address. Please enter a valid email.")
        else:
            # Place this where you define your interface components
            #rv_threshold = st.slider("Select RV % Est Threshold for Patient Care:", 100, 200, 150)
            standard = st.radio("Select Standard:", ('GLI', 'ECSC'),horizontal=True,index=None)
    
            if standard == 'GLI':
                has_fvc_pred = st.radio("Do You Have FVC % Predicted?", ('Yes', 'No'),horizontal=True,index=None)
                if has_fvc_pred == 'Yes':
                    gender = st.radio("Select Gender:", ('Male', 'Female'), horizontal=True,index=None)
            
                    if gender:
                        # Create layout with columns for age, measured FEV1, measured FVC, and FVC % Predicted
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            age = st.number_input("Enter Age (Years):", min_value=3, max_value=95, step=1, key='age_yes')
                        with col2:
                            measured_fev1 = st.number_input("Enter Measured FEV1 (X.XL):",min_value=0.0,format="%.2f", step=0.01, key='fev1_yes')
                        with col3:
                            measured_fvc = st.number_input("Enter Measured FVC (X.XL):", min_value=0.0, format="%.2f", step=0.01, key='fvc_yes')
                        with col4:
                            fvc_percent_predicted = st.number_input("Enter FVC % Predicted:", min_value=0.0,format="%.1f", step=0.1, key='fvc_percent_pred')
                            
                        # Store the button press result in a variable
                        evaluate_pressed = st.button('Evaluate')
                        
                        if evaluate_pressed and age and measured_fev1 and measured_fvc and fvc_percent_predicted:
                            measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
                            measured_fev1_fvc =  math.ceil(measured_fev1_fvc * 1000) / 1000
                            
                            # # Display measured values and provided FVC % Predicted
                            # col5, col6, col7, col8= st.columns(4)
                            # with col5:
                            #     st.metric(label="Measured FEV1", value=f"{measured_fev1:.2f} L")
                            # with col6:
                            #     st.metric(label="Measured FVC", value=f"{measured_fvc:.2f} L")
                            # with col7:
                            #     st.metric(label="FVC % Predicted", value=f"{fvc_percent_predicted:.1f}%")
                            # with col8:
                            #     # Calculate FEV1/FVC ratio from measured values and display
                            #     measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
                            #     measured_fev1_fvc =  math.ceil(measured_fev1_fvc * 1000) / 1000
                            #     st.metric(label="Measured FEV1/FVC", value=f"{measured_fev1_fvc:.3f}")
                            
                            col9, col10, col11,col12 = st.columns(4)
                            rv_percent_est = calculate_rv_est(fvc_percent_predicted, measured_fev1_fvc, age, gender)
                            # Calculate RV % Predicted Prevalence
                            RV150, RV175, RV200 = calculate_rv_predicted(rv_percent_est)
                            # Adding a row to display RV % est
                            col9, col10, col11, col12 = st.columns(4)
                            # with col9:
                            #     st.metric(label="RV % Estimate", value=f"{rv_percent_est:.1f}")
                            with col10:
                                st.metric(label="RV >150% Probability", value=f"{RV150:.1f}%")
                            # with col11:
                            #     st.metric(label="RV >175% Probability", value=f"{RV175:.1f}%")
                            # with col12:
                            #     st.metric(label="RV >200% Probability", value=f"{RV200:.1f}%")
        
                            # st.write("Final Result")
                            # # Final Result based on the RV% Est threshold
                            # if rv_percent_est >= rv_threshold:
                            #     st.success(f"Patient Can be Sent to Next Step 🟢")
                            # else:
                            #     st.error(f"Patient is Fit, No Further Care Required 🔴")
                            # Append the data to the CSV file in S3
                            append_to_s3(email, rv_threshold, standard, has_fvc_pred, gender, age, None, measured_fev1, measured_fvc, fvc_percent_predicted, None, None, None, rv_percent_est, RV150, None)
                        elif evaluate_pressed:
                            st.error("Please fill in all required fields before evaluating.")
                            
                elif has_fvc_pred == 'No':
                    gender = st.radio("Select Gender:", ('Male', 'Female'),horizontal=True,index=None)
    
                    if gender:
                        # Create a layout with columns for age, height, FEV1, and FVC on one line
                        col1, col2, col3, col4 = st.columns(4)
        
                        with col1:
                            age = st.number_input("Enter Age (Years):", min_value=3, max_value=95, step=1, key='age')
                        with col2:
                            height = st.number_input("Enter Height (in cm):", min_value=150.0, max_value=250.0, step=0.1, key='height',format="%.1f")
                        with col3:
                            measured_fev1 = st.number_input("Enter Measured FEV1 (X.XL):", min_value=0.0, format="%.2f", step=0.01, key='fev1')
                        with col4:
                            measured_fvc = st.number_input("Enter Measured FVC (X.XL):", min_value=0.0, format="%.2f", step=0.01, key='fvc')
    
                         # Store the button press result in a variable
                        calculate_pressed = st.button('Calculate')
                        
                        # Calculate button can be placed below the inputs or in a new line
                        if calculate_pressed and age and height and measured_fev1:
                            fev1, fvc, fev1_fvc = calculate_values(age, height, gender)
                            percent_predicted_fev1 = (measured_fev1 / fev1) * 100
                            percent_predicted_fvc = round(measured_fvc / fvc * 100,1) if fvc != 0 else 0
                            # Calculate FEV1/FVC ratio from measured values and display
                            measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
                            measured_fev1_fvc =  math.ceil(measured_fev1_fvc * 1000) / 1000
        
                            # col1, col2, col3, col4 = st.columns(4)  # Create four columns
                            
                            # with col1:
                            #     st.metric(label="Predicted FEV1", value=f"{fev1:.2f} L")
                            # with col2:
                            #     st.metric(label="Predicted FVC", value=f"{fvc:.2f} L")
                            # with col3:
                            #     st.metric(label="Predicted FEV1/FVC", value=f"{fev1_fvc:.2f}")
                            # with col4:
                            #     st.metric(label="% Predicted FVC", value=f"{percent_predicted_fvc:.1f}")
        
                            col5, col6, col7, col8 = st.columns(4)
                            
                            # with col5:
                            #     st.metric(label="Measured FEV1", value=f"{measured_fev1:.2f} L")
                            # with col6:
                            #     st.metric(label="Measured FVC", value=f"{measured_fvc:.2f} L")
                            # with col7:
                                # # Calculate FEV1/FVC ratio from measured values and display
                                # measured_fev1_fvc = measured_fev1 / measured_fvc if measured_fvc != 0 else 0
                                # measured_fev1_fvc =  math.ceil(measured_fev1_fvc * 1000) / 1000
                                # st.metric(label="Measured FEV1/FVC", value=f"{measured_fev1_fvc:.3f}")
                            rv_percent_est = calculate_rv_est(percent_predicted_fvc, measured_fev1_fvc, age, gender)
                            # Calculate RV % Predicted Prevalence
                            RV150, RV175, RV200 = calculate_rv_predicted(rv_percent_est)
                            # Adding a row to display RV % est
                            col9, col10, col11, col12 = st.columns(4)
                            # with col9:
                            #     st.metric(label="RV % Estimate", value=f"{rv_percent_est:.1f}")
                            with col10:
                                st.metric(label="RV >150% Probability", value=f"{RV150:.1f}%")
                            # with col11:
                            #     st.metric(label="RV >175% Probability", value=f"{RV175:.1f}%")
                            # with col12:
                            #     st.metric(label="RV >200% Probability", value=f"{RV200:.1f}%")
        
                            # st.write("Final Result")
                            # # Final Result based on the RV% Est threshold
                            # if rv_percent_est >= rv_threshold:
                            #     st.success(f"Patient Can be Sent to Next Step 🟢")
                            # else:
                            #     st.error(f"Patient is Fit, No Further Care Required 🔴")
                            # Append the data to the CSV file in S3
                            append_to_s3(email, rv_threshold, standard, has_fvc_pred, gender, age, height, measured_fev1, measured_fvc, percent_predicted_fvc, fev1, fvc, fev1_fvc, rv_percent_est, RV150, None)
    
                        elif calculate_pressed:
                            st.error("Please fill in all required fields before calculating.")
                            
            elif standard == 'ECSC':
                gender = st.radio("Select Sex:", (1, 0), format_func=lambda x: 'Male' if x == 1 else 'Female',index=None,horizontal=True)
                
                if gender is not None:  # Ensures that race is only shown if gender is selected
                    race = st.radio("Select Race:", (1, 2), format_func=lambda x: 'White' if x == 1 else 'Black',index=None,horizontal=True)
                    
                    if race is not None:  # Ensures that the remaining inputs are only shown if race is selected
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            age = st.number_input("Age:", min_value=3, max_value=95, step=1)
                        with col2:
                            height = st.number_input("Height (in cm):", min_value=100.0, max_value=250.0, step=0.1, format="%.1f")
                        with col3:
                            measured_fev1 = st.number_input("Enter Measured FEV1 (X.XL):", min_value=0.0, format="%.2f", step=0.01)
                        with col4:
                            measured_fvc = st.number_input("Enter Measured FVC (X.XL):", min_value=0.0, format="%.2f", step=0.01)
    
                        #Store the button press result in a variable
                        calculate_ECSC = st.button('Calculate ECSC')
    
                        if calculate_ECSC and age and height and measured_fev1 and measured_fvc:
                            pred_fvc = calculate_ecsc_fvc(age, height, measured_fev1, measured_fvc, gender, race)
                            fvc_percent_predicted_ecsc, fev1_fvc_ratio, rv_percent_est, rv150, rv175, rv200 = calculate_ecsc_metrics(age, height, measured_fev1, pred_fvc, measured_fvc,gender,race)
    
                            col8, col9, col10, col11 = st.columns(4)
                            # Display the calculated values
                            # with col5:
                            #     st.metric(label="Predicted FVC:", value=f"{pred_fvc}")
                            #     #st.write(f"Predicted FVC: {pred_fvc} L")
                            # with col6:
                            #     st.metric(label="FVC % Predicted:", value=f"{fvc_percent_predicted_ecsc}")
                            #     #st.write(f"FVC % Predicted: {fvc_percent_predicted_ecsc}%")
                            # with col7:
                            #     st.metric(label="FEV1/FVC Ratio:", value=f"{fev1_fvc_ratio}")
                            #     #st.write(f"FEV1/FVC Ratio: {fev1_fvc_ratio}%")
                            # with col8:    
                            #     st.metric(label="RV % Estimated:", value=f"{rv_percent_est}")
                            #     #st.write(f"RV % Estimated: {rv_percent_est}")
                            with col9:
                                #st.metric(label="RV >150% Probability", value=f"{rv150}%")
                                #st.write(f"RV >150% Probability: {rv150}%")
                                st.metric(label="RV >150% Probability", value=f"{rv150:.1f}%")
                            # with col10:
                            #     #st.metric(label="RV >175% Probability", value=f"{rv175}%")
                            #     #st.write(f"RV >175% Probability: {rv175}%")
                            #     st.metric(label="RV >175% Probability", value=f"{rv175:.1f}%")
                            # with col11:
                            #     #st.metric(label="RV >200% Probability", value=f"{rv200}%")
                            #     #st.write(f"RV >200% Probability: {rv200}%")
                            #     st.metric(label="RV >200% Probability", value=f"{rv200:.1f}%")
                        
                            # st.write("Final Result")
                            # # Final Result based on the RV% Est threshold
                            # if rv_percent_est >= rv_threshold:
                            #     st.success(f"Patient Can be Sent to Next Step 🟢")
                            # else:
                            #     st.error(f"Patient is Fit, No Further Care Required 🔴")
    
                            # Map numeric gender and race to descriptive text
                            gender_text = 'Male' if gender == 1 else 'Female'
                            race_text = 'White' if race == 1 else 'Black'
    
                            # Append the data to the CSV file in S3
                            append_to_s3(email, rv_threshold, standard, None, gender_text, age, height, measured_fev1, measured_fvc, fvc_percent_predicted_ecsc, None, None, None, rv_percent_est, rv150, race_text)
    
    
                        elif calculate_ECSC:
                            st.error("Please fill in all required fields before calculating.")
    else:
        st.write("Please enter an email address to continue.")
elif process_type == 'Batch':
    if email:
        is_email_valid = validate_email(email)
        if not is_email_valid:
            st.error("Invalid email address. Please enter a valid email.")
        else:
            standard = st.radio("Select Standard:", ('GLI', 'ECSC'), index=None)
            if standard == 'GLI':
                has_fvc_pred = st.radio("Do You Have FVC % Predicted?", ('Yes', 'No'),horizontal=True,index=None)
                if has_fvc_pred == 'Yes':
                    st.markdown("""
            #### Download Sample Excel Template
            
            To ensure your file has the correct format, download and use this [Download Excel](https://github.com/vinay-bhati/RV-Cal/raw/refs/heads/main/GLI_Has_FVC_Percent_Predicted.xlsx).
            """, unsafe_allow_html=True)
                    
                    st.markdown("""
            ### Batch Processing Instructions
            - **File Type:** Excel file (.xlsx)
            - **Required Columns:** Age, Gender, FEV1 (L), FVC (L), FVC % Predicted
            - **Data Format:**
              - **Age:** (3 - 95)
              - **Gender:** ('Male' or 'Female')
              - **FEV1:** ( x.xx, e.g., 2.34) liters (L)
              - **FVC:** ( x.xx, e.g., 3.45) liters (L)
              - **FVC % Predicted:** ( x.x, e.g., 12.4)
              
            Please ensure that your file adheres to the above format to avoid processing errors.
            """)

                    file = st.file_uploader("Upload Excel File", type=['xlsx'])
                    
                    # if file and st.button('Process Batch File'):
                    #     processed_data = process_gli_batch_excel(file)
                    #     if processed_data is not None and not processed_data.empty:
                    #         processed_data_csv = processed_data.to_csv(index=False).encode('utf-8')
                    #         st.download_button(
                    #             label="Download Processed Data",
                    #             data=processed_data_csv,
                    #             file_name='processed_data.csv',
                    #             mime='text/csv',
                    #         )
                    if file and st.button('Process Batch File'):
                        processed_data = process_gli_batch_excel(file)
                        
                        if processed_data is not None and not processed_data.empty:
                            # Create an Excel file in memory
                            output = io.BytesIO()
                            
                            # Write DataFrame to Excel
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                processed_data.to_excel(writer, index=False, sheet_name='Processed_Data')
                                writer.book.save(output)  # Save initial file
                            
                            # Load workbook to apply formatting
                            output.seek(0)
                            workbook = load_workbook(output)
                            worksheet = workbook['Processed_Data']
            
                            # Define a green fill color
                            green_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  # Light green
            
                            # List of columns to highlight
                            result_columns = ["FEV1/FVC", "Probability of RV>150"]
            
                            # Find column indexes dynamically
                            col_indexes = [processed_data.columns.get_loc(col) + 1 for col in result_columns]
            
                            # Apply fill to all rows in the selected columns
                            for col_idx in col_indexes:
                                for row in range(2, len(processed_data) + 2):  # Start from row 2 (headers in row 1)
                                    worksheet.cell(row=row, column=col_idx).fill = green_fill
            
                            # Save the updated workbook to memory
                            output = io.BytesIO()
                            workbook.save(output)
                            output.seek(0)
            
                            # Provide download button for Excel file
                            st.download_button(
                                label="Download Processed Data as Excel",
                                data=output,
                                file_name='processed_gli_data.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
            
                elif has_fvc_pred == 'No':
                    st.markdown("""
                    #### Download Sample Excel Template
                    
                    To help you prepare your data correctly, download and use this [Download Excel](https://github.com/vinay-bhati/RV-Cal/raw/refs/heads/main/GLI_No_FVC_Percent_Predicted.xlsx) template.
                    """, unsafe_allow_html=True)
                    
                    st.markdown("""
                    ### Batch Processing Instructions
                    - **File Type:** Excel file (.xlsx)
                    - **Required Columns:** Age, Gender, Height, FEV1 (L), FVC (L)
                    - **Data Format:**
                      - **Age:** (3 - 95)
                      - **Gender:**  ('Male' or 'Female')
                      - **Height:** (format x.x or x.xx, e.g., 175.5)
                      - **FEV1:** (format x.xx, e.g., 2.34) liters (L)
                      - **FVC:** ( x.xx, e.g., 3.45) liters (L)
                    
                    Please ensure that your file adheres to the above format to avoid processing errors.
                    """)

                    file = st.file_uploader("Upload Excel File", type=['xlsx'])
                    # if file and st.button('Process Batch File'):
                    #     processed_data = process_gli_batch_no_fvc_pred(file)
                    #     if processed_data is not None and not processed_data.empty:
                    #         processed_data_csv = processed_data.to_csv(index=False).encode('utf-8')
                    #         st.download_button(
                    #             label="Download Processed Data as CSV",
                    #             data=processed_data_csv,
                    #             file_name='processed_data.csv',
                    #             mime='text/csv'
                    #         )
                    if file and st.button('Process Batch File'):
                        processed_data = process_gli_batch_no_fvc_pred(file)
                        if processed_data is not None and not processed_data.empty:
                            # Create an Excel file in memory
                            output = io.BytesIO()
                    
                            # Write DataFrame to Excel
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                processed_data.to_excel(writer, index=False, sheet_name='Processed_Data')
                                writer.book.save(output)  # Save initial file
                            
                            # Load workbook to apply formatting
                            output.seek(0)
                            workbook = load_workbook(output)
                            worksheet = workbook['Processed_Data']
                    
                            # Define a green fill color
                            green_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  # Light green
                    
                            # List of columns to highlight
                            result_columns = ["FEV1/FVC", "FVC % Predicted", "Probability of RV>150"]
                    
                            # Find column indexes dynamically
                            col_indexes = [processed_data.columns.get_loc(col) + 1 for col in result_columns]
                    
                            # Apply fill to all rows in the selected columns
                            for col_idx in col_indexes:
                                for row in range(2, len(processed_data) + 2):  # Start from row 2 (headers in row 1)
                                    worksheet.cell(row=row, column=col_idx).fill = green_fill
                    
                            # Save the updated workbook to memory
                            output = io.BytesIO()
                            workbook.save(output)
                            output.seek(0)
                    
                            # Provide download button for Excel file
                            st.download_button(
                                label="Download Processed Data as Excel",
                                data=output,
                                file_name='processed_gli_no_fvc_pred.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
    
            elif standard == 'ECSC':
                st.markdown("""
                #### Download Sample Excel Template
                
                To help you prepare your data correctly, download and use this [Download Excel](https://github.com/vinay-bhati/RV-Cal/raw/refs/heads/main/Sample_ECSC_N.xlsx) template.
                """, unsafe_allow_html=True)
                
                st.markdown("""
                ### Batch Processing Instructions
                - **File Type:** Excel file (.xlsx)
                - **Required Columns:** Age, Gender, Height, FEV1 (L), FVC (L), Race
                - **Data Format:**
                  - **Age:** (3 - 95)
                  - **Gender:** ('Male' or 'Female')
                  - **Height:** (format x.x or x.xx, e.g., 175.5)
                  - **FEV1:** (format x.xx, e.g., 2.34) liters (L)
                  - **FVC:**  (format x.xx, e.g., 3.45) liters (L)
                  - **Race:**  ('White' or 'Black')
                
                Please ensure that your file adheres to the above format to avoid processing errors.
                """)
                
                file = st.file_uploader("Upload Excel File", type=['xlsx'])
                # if file and st.button('Process Batch File'):
                #     processed_data = process_ecsc_batch(file)
                #     if processed_data is not None and not processed_data.empty:
                #         processed_data_csv = processed_data.to_csv(index=False).encode('utf-8')
                #         st.download_button(
                #             label="Download Processed Data as CSV",
                #             data=processed_data_csv,
                #             file_name='processed_ecsc_data.csv',
                #             mime='text/csv'
                #         )
                # if file and st.button('Process Batch File'):
                #     processed_data = process_ecsc_batch(file)
                #     if processed_data is not None and not processed_data.empty:
                #         # Create an Excel file in memory
                #         output = io.BytesIO()
                #         with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                #             processed_data.to_excel(writer, index=False, sheet_name='Processed_Data')
                #             writer.close()
                #         output.seek(0)
                
                #         # Provide download button for Excel file
                #         st.download_button(
                #             label="Download Processed Data as Excel",
                #             data=output,
                #             file_name='processed_ecsc_data.xlsx',
                #             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                #         )
                if file and st.button('Process Batch File'):
                    processed_data = process_ecsc_batch(file)
                    if processed_data is not None and not processed_data.empty:
                        # Create an Excel file in memory
                        output = io.BytesIO()
                
                        # Write dataframe to Excel
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            processed_data.to_excel(writer, index=False, sheet_name='Processed_Data')
                            writer.book.save(output)  # Properly save the workbook
                
                        # Load workbook to apply formatting
                        output.seek(0)
                        workbook = load_workbook(output)
                        worksheet = workbook['Processed_Data']
                
                        # Define a green fill color
                        green_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  # Light green
                
                        # List of columns to highlight
                        result_columns = ["FVC % Predicted", "FEV1/FVC", "Probability of RV>150"]
                
                        # Find column indexes dynamically
                        col_indexes = [processed_data.columns.get_loc(col) + 1 for col in result_columns]
                
                        # Apply fill to all rows in the selected columns
                        for col_idx in col_indexes:
                            for row in range(2, len(processed_data) + 2):  # Start from row 2 (headers in row 1)
                                worksheet.cell(row=row, column=col_idx).fill = green_fill
                
                        # Save the updated workbook to memory
                        output = io.BytesIO()
                        workbook.save(output)
                        output.seek(0)
                
                        # Provide download button for Excel file
                        st.download_button(
                            label="Download Processed Data as Excel",
                            data=output,
                            file_name='processed_ecsc_data.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )

else:
        st.write("Please enter an email address to continue.")


